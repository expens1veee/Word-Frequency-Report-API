
import os
import sqlite3
import tempfile

import openpyxl
import pytest

from src.app.services.excel import generate_excel
from src.app.services.export import process_large_file_process


def _make_txt(content: str) -> str:
    """Создает временный .txt файл и возвращает его путь"""
    fd, path = tempfile.mkstemp(suffix=".txt")
    with os.fdopen(fd, "w", encoding="utf-8") as f:
        f.write(content)
    return path


class TestGenerateExcel:
    """Тесты потоковой записи в .xlsx."""

    def test_creates_file(self):
        """Функция должна создавать физический файл на диске"""
        path = generate_excel(iter([]))
        try:
            assert os.path.exists(path)
        finally:
            os.unlink(path)

    def test_file_is_valid_xlsx(self):
        """Созданный файл должен быть валидным xlsx-документом"""
        path = generate_excel(iter([]))
        try:
            wb = openpyxl.load_workbook(path)
            assert wb is not None
        finally:
            os.unlink(path)

    def test_header_row(self):
        """Первая строка должна содержать правильные заголовки"""
        path = generate_excel(iter([]))
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            row = [cell.value for cell in ws[1]]
            assert "Словоформа" in row
            assert "Кол-во во всём документе" in row
        finally:
            os.unlink(path)

    def test_data_rows_written(self):
        """Данные из генератора должны быть записаны в файл"""
        data = [
            ("дом", 5, "3,2"),
            ("кот", 1, "0,1"),
        ]
        path = generate_excel(iter(data))
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            # Строка 1 — заголовок, строка 2 — первые данные
            assert ws.cell(row=2, column=1).value == "дом"
            assert ws.cell(row=2, column=2).value == 5
            assert ws.cell(row=3, column=1).value == "кот"
        finally:
            os.unlink(path)

    def test_multicolumn_row_written(self):
        """Кортеж с несколькими чанками должен записываться корректно в несколько столбцов"""
        data = [
            ("слово", 10, "1,2,3 ...(продолжение ->)", "4,5,6"),
        ]
        path = generate_excel(iter(data))
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            row_values = [ws.cell(row=2, column=i).value for i in range(1, 5)]
            assert row_values[0] == "слово"
            assert row_values[1] == 10
            assert "продолжение" in row_values[2]
            assert row_values[3] == "4,5,6"
        finally:
            os.unlink(path)


class TestProcessLargeFileProcess:
    """Интеграционные тесты полного цикла: txt -> SQLite -> xlsx"""

    def test_returns_existing_xlsx(self, small_text_file):
        """Функция должна вернуть путь к реально существующему .xlsx файлу"""
        result_path = process_large_file_process(small_text_file)
        try:
            assert os.path.exists(result_path)
            assert result_path.endswith(".xlsx")
        finally:
            if os.path.exists(result_path):
                os.unlink(result_path)

    def test_output_is_valid_xlsx(self, small_text_file):
        """Результирующий файл должен быть валидным xlsx-документом"""
        result_path = process_large_file_process(small_text_file)
        try:
            wb = openpyxl.load_workbook(result_path)
            assert wb is not None
        finally:
            if os.path.exists(result_path):
                os.unlink(result_path)

    def test_output_has_header(self, small_text_file):
        """Первая строка xlsx должна содержать заголовки"""
        result_path = process_large_file_process(small_text_file)
        try:
            wb = openpyxl.load_workbook(result_path)
            ws = wb.active
            header = [ws.cell(row=1, column=i).value for i in range(1, 4)]
            assert header[0] == "Словоформа"
        finally:
            if os.path.exists(result_path):
                os.unlink(result_path)

    def test_words_are_lemmatized(self, small_text_file):
        """Слова в excel должны быть в нормальной форме (после лемматизации)"""
        result_path = process_large_file_process(small_text_file)
        try:
            wb = openpyxl.load_workbook(result_path)
            ws = wb.active
            words = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
            assert "житель" in words
            assert "жителем" not in words
            assert "жителю" not in words
        finally:
            if os.path.exists(result_path):
                os.unlink(result_path)

    def test_total_count_is_correct(self, small_text_file):
        """
        Столбец кол-во во всём документе должен содержать корректное 
        суммарное значение по всем строкам файла
        """
        result_path = process_large_file_process(small_text_file)
        try:
            wb = openpyxl.load_workbook(result_path)
            ws = wb.active
            total_by_word = {}
            for r in range(2, ws.max_row + 1):
                w = ws.cell(row=r, column=1).value
                c = ws.cell(row=r, column=2).value
                if w:
                    total_by_word[w] = c
            assert total_by_word.get("житель") == 3
        finally:
            if os.path.exists(result_path):
                os.unlink(result_path)

    def test_empty_file(self):
        """На пустом входном файле функция должна успешно вернуть xlsx только с заголовком"""
        path = _make_txt("")
        try:
            result_path = process_large_file_process(path)
            try:
                wb = openpyxl.load_workbook(result_path)
                ws = wb.active
                assert ws.max_row == 1
            finally:
                os.unlink(result_path)
        finally:
            os.unlink(path)

    def test_chunking_long_line(self):
        """
        Если частотное распределение по строкам превышает 32000 символов,
        оно должно быть разбито на несколько столбцов, а не обрезано
        """
        num_lines = 35000
        lines = ["яблоко"] + ["другое"] * (num_lines - 1)
        path = _make_txt("\n".join(lines))
        try:
            result_path = process_large_file_process(path)
            try:
                wb = openpyxl.load_workbook(result_path)
                ws = wb.active
                apple_row = None
                for r in range(2, ws.max_row + 1):
                    if ws.cell(row=r, column=1).value == "яблоко":
                        apple_row = r
                        break
                assert apple_row is not None, "яблоко не найдено в отчёте"
                # Должно быть заполнено хотя бы в 3 ячейке 
                cell_c = ws.cell(row=apple_row, column=3).value
                cell_d = ws.cell(row=apple_row, column=4).value
                assert cell_c is not None, "Первый чанк отсутствует"
                assert cell_d is not None, "Второй чанк отсутствует (нарезка не произошла)"
                # Первый чанк должен содержать метку о продолжении
                assert "продолжение" in str(cell_c)
            finally:
                os.unlink(result_path)
        finally:
            os.unlink(path)
